from __future__ import annotations

import re
import unicodedata
from difflib import SequenceMatcher
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill


ROOT = Path(r"D:\GAME\aksara-sunda")
SOURCE_GAME = ROOT / "rekapan-link2.xlsx"
CLASS_DIR = ROOT / "rekapan-link1-kelas"
OUTPUT = ROOT / "rekapan-pemain-game-kelas-dikoreksi.xlsx"

CLASS_LABELS = {
    "10.01": "10 Satu",
    "10.02": "10 Dua",
    "10.03": "10 Tiga",
    "10.04": "10 Empat",
    "10.05": "10 Lima",
    "10.06": "10 Enam",
    "10.07": "10 Tujuh",
    "10.08": "10 Delapan",
    "10.09": "10 Sembilan",
    "10.10": "10 Sepuluh",
    "10.11": "10 Sebelas",
    "10.12": "10 Dua Belas",
    "10.13": "10 Tiga Belas",
}


def normalize_name(value: object) -> str:
    text = "" if value is None else str(value)
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.lower()
    text = text.replace("`", "'")
    replacements = {
        "mochammad": "muhamad",
        "mohammad": "muhamad",
        "mohamad": "muhamad",
        "muhammad": "muhamad",
        "muhamad": "muhamad",
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def token_sort(value: str) -> str:
    return " ".join(sorted(value.split()))


def similarity(left: str, right: str) -> float:
    if not left or not right:
        return 0.0
    direct = SequenceMatcher(None, left, right).ratio()
    sorted_ratio = SequenceMatcher(None, token_sort(left), token_sort(right)).ratio()
    left_tokens = set(left.split())
    right_tokens = set(right.split())
    overlap = len(left_tokens & right_tokens) / max(1, len(left_tokens | right_tokens))
    return max(direct, sorted_ratio, overlap)


def class_code_from_filename(name: str) -> str | None:
    match = re.search(r"K-(10\.\d{2})", name)
    return match.group(1) if match else None


def load_roster() -> list[dict[str, object]]:
    roster: list[dict[str, object]] = []
    for path in sorted(CLASS_DIR.glob("*.xlsx")):
        code = class_code_from_filename(path.name)
        class_label = CLASS_LABELS.get(code or "")
        if not class_label:
            continue

        workbook = load_workbook(path, data_only=True)
        sheet = workbook[workbook.sheetnames[0]]
        for row in sheet.iter_rows(min_row=14, values_only=True):
            no, nis, name = row[:3]
            if not name:
                continue
            roster.append(
                {
                    "name": str(name).strip(),
                    "normalized": normalize_name(name),
                    "nis": nis,
                    "class": class_label,
                    "file": path.name,
                }
            )
    return roster


def find_match(name: str, roster: list[dict[str, object]], exact_index: dict[str, dict[str, object]]):
    normalized = normalize_name(name)
    if normalized in exact_index:
        return exact_index[normalized], 1.0, "exact"

    best = None
    best_score = 0.0
    for item in roster:
        score = similarity(normalized, str(item["normalized"]))
        if score > best_score:
            best = item
            best_score = score

    if best and best_score >= 0.86:
        return best, best_score, "fuzzy"
    return best, best_score, "unmatched"


def append_row(sheet, values):
    sheet.append(values)


def style_header(sheet):
    fill = PatternFill("solid", fgColor="E2F0D9")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = fill


def main() -> None:
    roster = load_roster()
    exact_index: dict[str, dict[str, object]] = {}
    for item in roster:
        normalized = str(item["normalized"])
        if normalized and normalized not in exact_index:
            exact_index[normalized] = item

    workbook = load_workbook(SOURCE_GAME)
    sheet = workbook[workbook.sheetnames[0]]

    headers = [cell.value for cell in sheet[1]]
    name_col = headers.index("Ngaran") + 1
    class_col = headers.index("Kelas") + 1

    correction_rows = []
    unmatched_rows = []

    for row_idx in range(2, sheet.max_row + 1):
        player_name = sheet.cell(row_idx, name_col).value
        chosen_class = sheet.cell(row_idx, class_col).value
        if not player_name:
            continue

        match, score, method = find_match(str(player_name), roster, exact_index)
        if method == "unmatched" or not match:
            unmatched_rows.append(
                [
                    player_name,
                    chosen_class,
                    getattr(match, "get", lambda *_: "")("name") if match else "",
                    getattr(match, "get", lambda *_: "")("class") if match else "",
                    round(score, 3),
                ]
            )
            continue

        official_class = match["class"]
        official_name = match["name"]
        if chosen_class != official_class:
            sheet.cell(row_idx, class_col).value = official_class
            correction_rows.append(
                [
                    player_name,
                    chosen_class,
                    official_class,
                    official_name,
                    match.get("nis", ""),
                    method,
                    round(score, 3),
                ]
            )

    for old_name in ["Koreksi Kelas", "Nama Belum Ketemu"]:
        if old_name in workbook.sheetnames:
            del workbook[old_name]

    corrections = workbook.create_sheet("Koreksi Kelas")
    append_row(
        corrections,
        [
            "Nama di game",
            "Kelas dipilih di game",
            "Kelas resmi",
            "Nama di daftar siswa",
            "NIS",
            "Metode cocok",
            "Skor cocok",
        ],
    )
    for row in correction_rows:
        append_row(corrections, row)
    style_header(corrections)

    unmatched = workbook.create_sheet("Nama Belum Ketemu")
    append_row(
        unmatched,
        [
            "Nama di game",
            "Kelas dipilih di game",
            "Kandidat nama terdekat",
            "Kandidat kelas",
            "Skor cocok",
        ],
    )
    for row in unmatched_rows:
        append_row(unmatched, row)
    style_header(unmatched)

    for ws in workbook.worksheets:
        for column in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in column)
            ws.column_dimensions[column[0].column_letter].width = min(max(max_len + 2, 12), 42)

    workbook.save(OUTPUT)

    print(f"Roster siswa: {len(roster)}")
    print(f"Kelas dikoreksi: {len(correction_rows)}")
    print(f"Nama belum ketemu: {len(unmatched_rows)}")
    print(OUTPUT)


if __name__ == "__main__":
    main()
