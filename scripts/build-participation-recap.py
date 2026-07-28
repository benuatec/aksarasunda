from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass, field
from datetime import datetime
from difflib import SequenceMatcher
from html.parser import HTMLParser
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(r"D:\GAME\aksara-sunda")
ROSTER_DIR = ROOT / "rekapan-link1-kelas"
SOURCE_DIR = ROOT / "rekap-linkB-sumber"
OUTPUT = ROOT / "rekap-kelengkapan-siswa-aksara-sunda.xlsx"
OUTPUT_COPY = Path(r"D:\GAME\aset-baru\rekap-kelengkapan-siswa-aksara-sunda.xlsx")

CLASS_CODES = [
    ("10.01", "10 Satu"),
    ("10.02", "10 Dua"),
    ("10.03", "10 Tiga"),
    ("10.04", "10 Empat"),
    ("10.05", "10 Lima"),
    ("10.06", "10 Enam"),
    ("10.07", "10 Tujuh"),
    ("10.08", "10 Delapan"),
    ("10.09", "10 Sembilan"),
    ("10.10", "10 Sepuluh"),
    ("10.11", "10 Sebelas"),
    ("10.12", "10 Dua Belas"),
    ("10.13", "10 Tiga Belas"),
]

CLASS_BY_CODE = dict(CLASS_CODES)
CODE_BY_CLASS = {label: code for code, label in CLASS_CODES}
CLASS_FROM_FORM = {
    "x.1": "10 Satu",
    "x.01": "10 Satu",
    "x.2": "10 Dua",
    "x.02": "10 Dua",
    "x.3": "10 Tiga",
    "x.03": "10 Tiga",
    "x.4": "10 Empat",
    "x.04": "10 Empat",
    "x.5": "10 Lima",
    "x.05": "10 Lima",
    "x.6": "10 Enam",
    "x.06": "10 Enam",
    "x.7": "10 Tujuh",
    "x.07": "10 Tujuh",
    "x.8": "10 Delapan",
    "x.08": "10 Delapan",
    "x.9": "10 Sembilan",
    "x.09": "10 Sembilan",
    "x.10": "10 Sepuluh",
    "x.11": "10 Sebelas",
    "x.12": "10 Dua Belas",
    "x.13": "10 Tiga Belas",
}


@dataclass
class Student:
    key: str
    no: Any
    nis: Any
    name: str
    class_code: str
    class_name: str
    norm: str
    pre: dict[str, Any] | None = None
    post: dict[str, Any] | None = None
    questionnaire: dict[str, Any] | None = None
    game: dict[str, Any] | None = None
    notes: list[str] = field(default_factory=list)


def normalize_name(value: Any) -> str:
    text = "" if value is None else str(value)
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.lower().replace("`", "'")
    replacements = {
        "mochammad": "muhamad",
        "mohammad": "muhamad",
        "mohamad": "muhamad",
        "muhammad": "muhamad",
        "m.": "muhamad ",
        "m ": "muhamad ",
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def token_sort(value: str) -> str:
    return " ".join(sorted(value.split()))


def similarity(left: str, right: str) -> float:
    if not left or not right:
        return 0.0
    direct = SequenceMatcher(None, left, right).ratio()
    sorted_ratio = SequenceMatcher(None, token_sort(left), token_sort(right)).ratio()
    left_tokens = set(left.split())
    right_tokens = set(right.split())
    overlap = len(left_tokens & right_tokens) / max(1, len(left_tokens | right_tokens))
    containment = 0.0
    if len(left) >= 5 and (left in right or right in left):
        containment = min(len(left), len(right)) / max(len(left), len(right))
    return max(direct, sorted_ratio, overlap, containment)


def class_code_from_filename(name: str) -> str | None:
    match = re.search(r"K-(10\.\d{2})", name)
    return match.group(1) if match else None


def parse_class_hint(value: Any, rombel: Any = None) -> str | None:
    raw = "" if value is None else str(value).strip().lower()
    raw = raw.replace(" ", "")
    if raw in CLASS_FROM_FORM:
        return CLASS_FROM_FORM[raw]
    if raw in {"x", "10"} and rombel is not None:
        try:
            return CLASS_BY_CODE.get(f"10.{int(float(rombel)):02d}")
        except (TypeError, ValueError):
            return None
    if raw.startswith("10."):
        return CLASS_BY_CODE.get(raw)
    return None


def load_roster() -> list[Student]:
    students: list[Student] = []
    for path in sorted(ROSTER_DIR.glob("*.xlsx")):
        code = class_code_from_filename(path.name)
        class_name = CLASS_BY_CODE.get(code or "")
        if not code or not class_name:
            continue
        workbook = load_workbook(path, data_only=True)
        sheet = workbook[workbook.sheetnames[0]]
        for row in sheet.iter_rows(min_row=14, values_only=True):
            no, nis, name = row[:3]
            if not name:
                continue
            clean_name = str(name).strip()
            students.append(
                Student(
                    key=f"{code}-{nis}-{normalize_name(clean_name)}",
                    no=no,
                    nis=nis,
                    name=clean_name,
                    class_code=code,
                    class_name=class_name,
                    norm=normalize_name(clean_name),
                )
            )
    return students


def build_exact_index(students: list[Student]) -> dict[str, list[Student]]:
    exact: dict[str, list[Student]] = {}
    for student in students:
        exact.setdefault(student.norm, []).append(student)
    return exact


def find_student(
    name: Any,
    students: list[Student],
    exact_index: dict[str, list[Student]],
    class_hint: str | None = None,
) -> tuple[Student | None, float, str]:
    norm = normalize_name(name)
    if not norm:
        return None, 0.0, "empty"

    exact_matches = exact_index.get(norm, [])
    if class_hint:
        exact_class = [item for item in exact_matches if item.class_name == class_hint]
        if len(exact_class) == 1:
            return exact_class[0], 1.0, "exact+class"
    if len(exact_matches) == 1:
        return exact_matches[0], 1.0, "exact"

    pool = [item for item in students if item.class_name == class_hint] if class_hint else students
    best = None
    best_score = 0.0
    for item in pool:
        score = similarity(norm, item.norm)
        if score > best_score:
            best = item
            best_score = score

    threshold = 0.76 if class_hint else 0.86
    if best and best_score >= threshold:
        return best, best_score, "fuzzy+class" if class_hint else "fuzzy"
    return best, best_score, "unmatched"


def timestamp_value(value: Any) -> datetime:
    if isinstance(value, datetime):
        return value
    return datetime.min


def is_better_attempt(new: dict[str, Any], old: dict[str, Any] | None, score_key: str = "score") -> bool:
    if old is None:
        return True
    new_score = new.get(score_key)
    old_score = old.get(score_key)
    try:
        if float(new_score) != float(old_score):
            return float(new_score) > float(old_score)
    except (TypeError, ValueError):
        pass
    return timestamp_value(new.get("timestamp")) > timestamp_value(old.get("timestamp"))


def load_form_records(path: Path, source: str) -> pd.DataFrame:
    return pd.read_excel(path, sheet_name=0)


def assign_pre_post(
    df: pd.DataFrame,
    students: list[Student],
    exact_index: dict[str, list[Student]],
    source: str,
    unmatched: list[list[Any]],
) -> None:
    name_col = "Nama Lengkap"
    class_col = "Kelas"
    score_col = "Score"
    for _, row in df.iterrows():
        class_hint = parse_class_hint(row.get(class_col), row.get("Rombel"))
        student, score, method = find_student(row.get(name_col), students, exact_index, class_hint)
        record = {
            "timestamp": row.get("Timestamp"),
            "score": row.get(score_col),
            "class_input": row.get(class_col),
            "method": method,
            "match_score": round(score, 3),
        }
        if not student:
            unmatched.append([source, row.get(name_col), row.get(class_col), "", "", round(score, 3), method])
            continue
        if method == "unmatched":
            unmatched.append([source, row.get(name_col), row.get(class_col), student.name, student.class_name, round(score, 3), method])
            continue
        attr = "pre" if source == "Pre-test" else "post"
        if is_better_attempt(record, getattr(student, attr)):
            setattr(student, attr, record)


def assign_questionnaire(
    df: pd.DataFrame,
    students: list[Student],
    exact_index: dict[str, list[Student]],
    unmatched: list[list[Any]],
) -> None:
    question_cols = [col for col in df.columns if col not in {"Timestamp", "Nama Lengkap", "Usia", "Jenis kelamin", "Kelas"}]
    for _, row in df.iterrows():
        class_hint = parse_class_hint(row.get("Kelas"))
        student, score, method = find_student(row.get("Nama Lengkap"), students, exact_index, class_hint)
        values = [row.get(col) for col in question_cols]
        numeric = [float(v) for v in values if pd.notna(v)]
        record = {
            "timestamp": row.get("Timestamp"),
            "average": round(sum(numeric) / len(numeric), 2) if numeric else None,
            "total": sum(numeric) if numeric else None,
            "class_input": row.get("Kelas"),
            "method": method,
            "match_score": round(score, 3),
        }
        if not student:
            unmatched.append(["Kuesioner", row.get("Nama Lengkap"), row.get("Kelas"), "", "", round(score, 3), method])
            continue
        if method == "unmatched":
            unmatched.append(["Kuesioner", row.get("Nama Lengkap"), row.get("Kelas"), student.name, student.class_name, round(score, 3), method])
            continue
        if is_better_attempt(record, student.questionnaire, "average"):
            student.questionnaire = record


def assign_game(
    game_path: Path,
    students: list[Student],
    exact_index: dict[str, list[Student]],
    unmatched: list[list[Any]],
) -> None:
    df = pd.DataFrame(parse_first_html_table(game_path))
    for _, row in df.iterrows():
        class_hint = str(row.get("Kelas")) if pd.notna(row.get("Kelas")) else None
        student, score, method = find_student(row.get("Ngaran"), students, exact_index, class_hint)
        record = {
            "level": row.get("Level Pangluhurna"),
            "score": row.get("Jumlah Skor"),
            "plays": row.get("Jumlah Maén"),
            "status": row.get("Status"),
            "class_input": row.get("Kelas"),
            "method": method,
            "match_score": round(score, 3),
        }
        if not student:
            unmatched.append(["Game", row.get("Ngaran"), row.get("Kelas"), "", "", round(score, 3), method])
            continue
        if method == "unmatched":
            unmatched.append(["Game", row.get("Ngaran"), row.get("Kelas"), student.name, student.class_name, round(score, 3), method])
            continue
        if is_better_attempt(record, student.game, "score"):
            student.game = record


class SimpleTableParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.rows: list[list[str]] = []
        self.current_row: list[str] | None = None
        self.current_cell: list[str] | None = None
        self.in_cell = False

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        if tag == "tr":
            self.current_row = []
        elif tag in {"td", "th"} and self.current_row is not None:
            self.current_cell = []
            self.in_cell = True

    def handle_data(self, data: str) -> None:
        if self.in_cell and self.current_cell is not None:
            self.current_cell.append(data)

    def handle_endtag(self, tag: str) -> None:
        if tag in {"td", "th"} and self.current_row is not None and self.current_cell is not None:
            self.current_row.append("".join(self.current_cell).strip())
            self.current_cell = None
            self.in_cell = False
        elif tag == "tr" and self.current_row is not None:
            if self.current_row:
                self.rows.append(self.current_row)
            self.current_row = None


def parse_first_html_table(path: Path) -> list[dict[str, Any]]:
    parser = SimpleTableParser()
    parser.feed(path.read_text(encoding="utf-8", errors="ignore"))
    if not parser.rows:
        return []
    headers = parser.rows[0]
    return [dict(zip(headers, row)) for row in parser.rows[1:] if any(row)]


def yesno(value: Any) -> str:
    return "Sudah" if value else "Belum"


def style_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="166534")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for column in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in column)
        ws.column_dimensions[get_column_letter(column[0].column)].width = min(max(max_len + 2, 10), 32)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def append_student_rows(ws, rows: list[Student]) -> None:
    ws.append(
        [
            "No",
            "NIS",
            "Nama Resmi",
            "Kelas Resmi",
            "Pre-test",
            "Nilai Pre",
            "Post-test",
            "Nilai Post",
            "Kuesioner",
            "Rata-rata Kuesioner",
            "Main Game",
            "Level Game",
            "Skor Game",
            "Jumlah Main",
            "Lengkap Semua",
            "Catatan",
        ]
    )
    for student in rows:
        complete = all([student.pre, student.post, student.questionnaire, student.game])
        note_parts = []
        for label, record in [("pre", student.pre), ("post", student.post), ("kuesioner", student.questionnaire), ("game", student.game)]:
            if record and record.get("class_input") and str(record.get("class_input")) != student.class_name:
                note_parts.append(f"{label}: input kelas {record.get('class_input')}")
        ws.append(
            [
                student.no,
                student.nis,
                student.name,
                student.class_name,
                yesno(student.pre),
                student.pre.get("score") if student.pre else "",
                yesno(student.post),
                student.post.get("score") if student.post else "",
                yesno(student.questionnaire),
                student.questionnaire.get("average") if student.questionnaire else "",
                yesno(student.game),
                student.game.get("level") if student.game else "",
                student.game.get("score") if student.game else "",
                student.game.get("plays") if student.game else "",
                "Ya" if complete else "Belum",
                "; ".join(note_parts),
            ]
        )


def main() -> None:
    students = load_roster()
    exact_index = build_exact_index(students)
    unmatched: list[list[Any]] = []

    assign_pre_post(load_form_records(SOURCE_DIR / "pre-test-jawaban.xlsx", "Pre-test"), students, exact_index, "Pre-test", unmatched)
    assign_pre_post(load_form_records(SOURCE_DIR / "post-test-jawaban.xlsx", "Post-test"), students, exact_index, "Post-test", unmatched)
    assign_questionnaire(load_form_records(SOURCE_DIR / "kuesioner-jawaban.xlsx", "Kuesioner"), students, exact_index, unmatched)
    assign_game(SOURCE_DIR / "ringkasan-game.xlsx", students, exact_index, unmatched)

    workbook = Workbook()
    summary = workbook.active
    summary.title = "Ringkasan"
    summary.append(
        [
            "Kelas",
            "Jumlah Siswa",
            "Sudah Pre-test",
            "Sudah Post-test",
            "Sudah Kuesioner",
            "Sudah Main Game",
            "Lengkap Semua",
            "Belum Lengkap",
        ]
    )

    for code, class_name in CLASS_CODES:
        class_students = [student for student in students if student.class_code == code]
        summary.append(
            [
                class_name,
                len(class_students),
                sum(1 for item in class_students if item.pre),
                sum(1 for item in class_students if item.post),
                sum(1 for item in class_students if item.questionnaire),
                sum(1 for item in class_students if item.game),
                sum(1 for item in class_students if all([item.pre, item.post, item.questionnaire, item.game])),
                sum(1 for item in class_students if not all([item.pre, item.post, item.questionnaire, item.game])),
            ]
        )
        ws = workbook.create_sheet(code)
        append_student_rows(ws, class_students)
        style_sheet(ws)

    unmatched_ws = workbook.create_sheet("Data Tidak Cocok")
    unmatched_ws.append(["Sumber", "Nama Input", "Kelas Input", "Kandidat Nama Resmi", "Kandidat Kelas", "Skor Cocok", "Metode"])
    for row in unmatched:
        unmatched_ws.append(row)

    style_sheet(summary)
    style_sheet(unmatched_ws)
    workbook.save(OUTPUT)
    OUTPUT_COPY.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(OUTPUT_COPY)

    print(f"Jumlah siswa resmi: {len(students)}")
    print(f"Output: {OUTPUT}")
    print(f"Copy: {OUTPUT_COPY}")
    print(f"Data input belum cocok: {len(unmatched)}")


if __name__ == "__main__":
    main()
