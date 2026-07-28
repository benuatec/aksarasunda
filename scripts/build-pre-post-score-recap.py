from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass, field
from datetime import datetime
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(r"D:\GAME\aksara-sunda")
ROSTER_DIR = ROOT / "rekapan-link1-kelas"
SOURCE_DIR = ROOT / "rekap-linkB-sumber"
OUTPUT = ROOT / "rekap-nilai-pre-post-siswa.xlsx"
OUTPUT_COPY = Path(r"D:\GAME\aset-baru\rekap-nilai-pre-post-siswa.xlsx")

CLASS_CODES = [
    ("10.01", "10 Satu"),
    ("10.02", "10 Dua"),
    ("10.03", "10 Tiga"),
    ("10.04", "10 Empat"),
    ("10.05", "10 Lima"),
    ("10.06", "10 Enam"),
    ("10.07", "10 Tujuh"),
    ("10.08", "10 Delapan"),
    ("10.09", "10 Sembilan"),
    ("10.10", "10 Sepuluh"),
    ("10.11", "10 Sebelas"),
    ("10.12", "10 Dua Belas"),
    ("10.13", "10 Tiga Belas"),
]

CLASS_BY_CODE = dict(CLASS_CODES)
CLASS_FROM_FORM = {
    "x.1": "10 Satu",
    "x.01": "10 Satu",
    "x1": "10 Satu",
    "x.2": "10 Dua",
    "x.02": "10 Dua",
    "x2": "10 Dua",
    "x.3": "10 Tiga",
    "x.03": "10 Tiga",
    "x3": "10 Tiga",
    "x.4": "10 Empat",
    "x.04": "10 Empat",
    "x4": "10 Empat",
    "x.5": "10 Lima",
    "x.05": "10 Lima",
    "x5": "10 Lima",
    "x.6": "10 Enam",
    "x.06": "10 Enam",
    "x6": "10 Enam",
    "x.7": "10 Tujuh",
    "x.07": "10 Tujuh",
    "x7": "10 Tujuh",
    "x.8": "10 Delapan",
    "x.08": "10 Delapan",
    "x8": "10 Delapan",
    "x.9": "10 Sembilan",
    "x.09": "10 Sembilan",
    "x9": "10 Sembilan",
    "x.10": "10 Sepuluh",
    "x10": "10 Sepuluh",
    "x.11": "10 Sebelas",
    "x11": "10 Sebelas",
    "x.12": "10 Dua Belas",
    "x12": "10 Dua Belas",
    "x.13": "10 Tiga Belas",
    "x13": "10 Tiga Belas",
}

MANUAL_NAME_ALIASES = {
    "aila": "aila artika",
    "jayyinatul": "jayyinatul assiyah jaami",
    "moch assyakir muhamad a": "moch assyakir maulidi abdillah",
    "zaki ibrahim mr": "zaky ibrahim muhamad rifay",
    "zaki ibrahim": "zaky ibrahim muhamad rifay",
    "arilmaulana": "aril maulana ridwan",
    "naima rahma rm": "naima rahma raudhatul muna",
    "naima rahma r m": "naima rahma raudhatul muna",
    "nizam": "nizam mulkil akbar",
    "farhan a muhamad": "farhan ainul mubarok",
    "farhan a m": "farhan ainul mubarok",
    "karina": "karina melda septiara",
}


@dataclass
class Attempt:
    source: str
    timestamp: Any
    score: Any
    name_input: Any
    class_input: Any
    method: str
    match_score: float


@dataclass
class Student:
    key: str
    no: Any
    nis: Any
    name: str
    class_code: str
    class_name: str
    norm: str
    pre_attempts: list[Attempt] = field(default_factory=list)
    post_attempts: list[Attempt] = field(default_factory=list)


def normalize_name(value: Any) -> str:
    text = "" if value is None else str(value)
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.lower().replace("`", "'")
    replacements = {
        "mochammad": "muhamad",
        "mohammad": "muhamad",
        "mohamad": "muhamad",
        "muhammad": "muhamad",
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    text = re.sub(r"\bm\.\s*", "muhamad ", text)
    text = re.sub(r"\bm\s+", "muhamad ", text)
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def token_sort(value: str) -> str:
    return " ".join(sorted(value.split()))


def similarity(left: str, right: str) -> float:
    if not left or not right:
        return 0.0
    direct = SequenceMatcher(None, left, right).ratio()
    sorted_ratio = SequenceMatcher(None, token_sort(left), token_sort(right)).ratio()
    left_tokens = set(left.split())
    right_tokens = set(right.split())
    overlap = len(left_tokens & right_tokens) / max(1, len(left_tokens | right_tokens))
    subset_bonus = 0.0
    if min(len(left_tokens), len(right_tokens)) >= 2 and (left_tokens <= right_tokens or right_tokens <= left_tokens):
        subset_bonus = 0.92
    containment = 0.0
    if len(left) >= 5 and (left in right or right in left):
        containment = min(len(left), len(right)) / max(len(left), len(right))
    return max(direct, sorted_ratio, overlap, containment, subset_bonus)


def class_code_from_filename(name: str) -> str | None:
    match = re.search(r"K-(10\.\d{2})", name)
    return match.group(1) if match else None


def parse_class_hint(value: Any, rombel: Any = None) -> str | None:
    raw = "" if value is None else str(value).strip().lower().replace(" ", "")
    if raw in CLASS_FROM_FORM:
        return CLASS_FROM_FORM[raw]
    if raw in {"x", "10"} and rombel is not None:
        try:
            return CLASS_BY_CODE.get(f"10.{int(float(rombel)):02d}")
        except (TypeError, ValueError):
            return None
    if raw.startswith("10."):
        return CLASS_BY_CODE.get(raw)
    return None


def is_outside_target_class(value: Any) -> bool:
    raw = "" if value is None else str(value).strip().lower().replace(" ", "")
    return raw.startswith("xi")


def load_roster() -> list[Student]:
    students: list[Student] = []
    for path in sorted(ROSTER_DIR.glob("*.xlsx")):
        code = class_code_from_filename(path.name)
        class_name = CLASS_BY_CODE.get(code or "")
        if not code or not class_name:
            continue
        workbook = load_workbook(path, data_only=True)
        sheet = workbook[workbook.sheetnames[0]]
        for row in sheet.iter_rows(min_row=14, values_only=True):
            no, nis, name = row[:3]
            if not name:
                continue
            clean_name = str(name).strip()
            students.append(
                Student(
                    key=f"{code}-{nis}-{normalize_name(clean_name)}",
                    no=no,
                    nis=nis,
                    name=clean_name,
                    class_code=code,
                    class_name=class_name,
                    norm=normalize_name(clean_name),
                )
            )
    return students


def build_exact_index(students: list[Student]) -> dict[str, list[Student]]:
    exact: dict[str, list[Student]] = {}
    for student in students:
        exact.setdefault(student.norm, []).append(student)
    return exact


def find_student(
    name: Any,
    students: list[Student],
    exact_index: dict[str, list[Student]],
    class_hint: str | None = None,
) -> tuple[Student | None, float, str]:
    norm = normalize_name(name)
    if not norm:
        return None, 0.0, "empty"

    alias_norm = MANUAL_NAME_ALIASES.get(norm)
    if alias_norm:
        alias_matches = exact_index.get(alias_norm, [])
        if class_hint:
            alias_matches = [item for item in alias_matches if item.class_name == class_hint]
        if len(alias_matches) == 1:
            return alias_matches[0], 1.0, "alias"

    exact_matches = exact_index.get(norm, [])
    if class_hint:
        exact_class = [item for item in exact_matches if item.class_name == class_hint]
        if len(exact_class) == 1:
            return exact_class[0], 1.0, "exact+class"
    if len(exact_matches) == 1:
        return exact_matches[0], 1.0, "exact"

    pool = [item for item in students if item.class_name == class_hint] if class_hint else students
    best = None
    best_score = 0.0
    for item in pool:
        score = similarity(norm, item.norm)
        if score > best_score:
            best = item
            best_score = score

    threshold = 0.76 if class_hint else 0.86
    if best and best_score >= threshold:
        return best, best_score, "fuzzy+class" if class_hint else "fuzzy"
    return best, best_score, "unmatched"


def timestamp_key(value: Any) -> datetime:
    if isinstance(value, datetime):
        return value
    try:
        parsed = pd.to_datetime(value)
        if pd.isna(parsed):
            return datetime.max
        return parsed.to_pydatetime()
    except (TypeError, ValueError):
        return datetime.max


def format_timestamp(value: Any) -> str:
    stamp = timestamp_key(value)
    if stamp == datetime.max:
        return ""
    return stamp.strftime("%Y-%m-%d %H:%M:%S")


def earliest(attempts: list[Attempt]) -> Attempt | None:
    if not attempts:
        return None
    return sorted(attempts, key=lambda item: timestamp_key(item.timestamp))[0]


def load_attempts(
    path: Path,
    source: str,
    students: list[Student],
    exact_index: dict[str, list[Student]],
    unmatched: list[list[Any]],
    ignored: list[list[Any]],
) -> None:
    df = pd.read_excel(path, sheet_name=0)
    for _, row in df.iterrows():
        class_hint = parse_class_hint(row.get("Kelas"), row.get("Rombel"))
        if not class_hint and is_outside_target_class(row.get("Kelas")):
            ignored.append(
                [
                    source,
                    row.get("Nama Lengkap"),
                    row.get("Kelas"),
                    row.get("Score"),
                    format_timestamp(row.get("Timestamp")),
                    "Diabaikan karena kelas input di luar kelas 10",
                ]
            )
            continue
        student, score, method = find_student(row.get("Nama Lengkap"), students, exact_index, class_hint)
        attempt = Attempt(
            source=source,
            timestamp=row.get("Timestamp"),
            score=row.get("Score"),
            name_input=row.get("Nama Lengkap"),
            class_input=row.get("Kelas"),
            method=method,
            match_score=round(score, 3),
        )
        if not student:
            unmatched.append([source, attempt.name_input, attempt.class_input, "", "", attempt.match_score, method, attempt.score, format_timestamp(attempt.timestamp)])
            continue
        if method == "unmatched":
            unmatched.append([source, attempt.name_input, attempt.class_input, student.name, student.class_name, attempt.match_score, method, attempt.score, format_timestamp(attempt.timestamp)])
            continue
        if source == "Pre-test":
            student.pre_attempts.append(attempt)
        else:
            student.post_attempts.append(attempt)


def style_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="166534")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for column in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in column)
        ws.column_dimensions[get_column_letter(column[0].column)].width = min(max(max_len + 2, 10), 34)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def append_class_sheet(ws, rows: list[Student]) -> None:
    ws.append(
        [
            "No",
            "NIS",
            "Nama Resmi",
            "Kelas Resmi",
            "Nilai Pre-test",
            "Waktu Pre-test",
            "Input Kelas Pre",
            "Jumlah Isi Pre",
            "Nilai Post-test",
            "Waktu Post-test",
            "Input Kelas Post",
            "Jumlah Isi Post",
            "Selisih Post-Pre",
            "Catatan",
        ]
    )
    for student in rows:
        pre = earliest(student.pre_attempts)
        post = earliest(student.post_attempts)
        note_parts = []
        if pre and pre.class_input and str(pre.class_input) != student.class_name:
            note_parts.append(f"Pre input kelas: {pre.class_input}")
        if post and post.class_input and str(post.class_input) != student.class_name:
            note_parts.append(f"Post input kelas: {post.class_input}")
        if len(student.pre_attempts) > 1:
            note_parts.append("Pre ganda, dipilih waktu paling awal")
        if len(student.post_attempts) > 1:
            note_parts.append("Post ganda, dipilih waktu paling awal")

        delta = ""
        try:
            if pre and post:
                delta = float(post.score) - float(pre.score)
        except (TypeError, ValueError):
            delta = ""

        ws.append(
            [
                student.no,
                student.nis,
                student.name,
                student.class_name,
                pre.score if pre else "",
                format_timestamp(pre.timestamp) if pre else "",
                pre.class_input if pre else "",
                len(student.pre_attempts),
                post.score if post else "",
                format_timestamp(post.timestamp) if post else "",
                post.class_input if post else "",
                len(student.post_attempts),
                delta,
                "; ".join(note_parts),
            ]
        )


def append_duplicate_sheet(ws, students: list[Student]) -> None:
    ws.append(["Sumber", "Nama Resmi", "Kelas Resmi", "Jumlah Isi", "Nilai Dipakai", "Waktu Dipakai", "Semua Pengisian"])
    for student in students:
        for source, attempts in [("Pre-test", student.pre_attempts), ("Post-test", student.post_attempts)]:
            if len(attempts) <= 1:
                continue
            selected = earliest(attempts)
            details = "; ".join(
                f"{format_timestamp(attempt.timestamp)} nilai {attempt.score} input kelas {attempt.class_input}"
                for attempt in sorted(attempts, key=lambda item: timestamp_key(item.timestamp))
            )
            ws.append(
                [
                    source,
                    student.name,
                    student.class_name,
                    len(attempts),
                    selected.score if selected else "",
                    format_timestamp(selected.timestamp) if selected else "",
                    details,
                ]
            )


def main() -> None:
    students = load_roster()
    exact_index = build_exact_index(students)
    unmatched: list[list[Any]] = []
    ignored: list[list[Any]] = []

    load_attempts(SOURCE_DIR / "pre-test-jawaban.xlsx", "Pre-test", students, exact_index, unmatched, ignored)
    load_attempts(SOURCE_DIR / "post-test-jawaban.xlsx", "Post-test", students, exact_index, unmatched, ignored)

    workbook = Workbook()
    summary = workbook.active
    summary.title = "Ringkasan"
    summary.append(
        [
            "Kelas",
            "Jumlah Siswa",
            "Ada Nilai Pre",
            "Ada Nilai Post",
            "Ada Keduanya",
            "Rata-rata Pre",
            "Rata-rata Post",
            "Rata-rata Selisih",
            "Pre Ganda",
            "Post Ganda",
        ]
    )

    for code, class_name in CLASS_CODES:
        class_students = [student for student in students if student.class_code == code]
        pre_scores = [float(pre.score) for student in class_students if (pre := earliest(student.pre_attempts)) and pd.notna(pre.score)]
        post_scores = [float(post.score) for student in class_students if (post := earliest(student.post_attempts)) and pd.notna(post.score)]
        deltas = []
        for student in class_students:
            pre = earliest(student.pre_attempts)
            post = earliest(student.post_attempts)
            if pre and post and pd.notna(pre.score) and pd.notna(post.score):
                deltas.append(float(post.score) - float(pre.score))

        summary.append(
            [
                class_name,
                len(class_students),
                sum(1 for item in class_students if item.pre_attempts),
                sum(1 for item in class_students if item.post_attempts),
                sum(1 for item in class_students if item.pre_attempts and item.post_attempts),
                round(sum(pre_scores) / len(pre_scores), 2) if pre_scores else "",
                round(sum(post_scores) / len(post_scores), 2) if post_scores else "",
                round(sum(deltas) / len(deltas), 2) if deltas else "",
                sum(1 for item in class_students if len(item.pre_attempts) > 1),
                sum(1 for item in class_students if len(item.post_attempts) > 1),
            ]
        )

        ws = workbook.create_sheet(code)
        append_class_sheet(ws, class_students)
        style_sheet(ws)

    duplicate_ws = workbook.create_sheet("Pengisian Ganda")
    append_duplicate_sheet(duplicate_ws, students)

    unmatched_ws = workbook.create_sheet("Data Tidak Cocok")
    unmatched_ws.append(["Sumber", "Nama Input", "Kelas Input", "Kandidat Nama Resmi", "Kandidat Kelas", "Skor Cocok", "Metode", "Nilai", "Waktu"])
    for row in unmatched:
        unmatched_ws.append(row)

    ignored_ws = workbook.create_sheet("Diabaikan")
    ignored_ws.append(["Sumber", "Nama Input", "Kelas Input", "Nilai", "Waktu", "Alasan"])
    for row in ignored:
        ignored_ws.append(row)

    style_sheet(summary)
    style_sheet(duplicate_ws)
    style_sheet(unmatched_ws)
    style_sheet(ignored_ws)

    workbook.save(OUTPUT)
    OUTPUT_COPY.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(OUTPUT_COPY)

    print(f"Jumlah siswa resmi: {len(students)}")
    print(f"Output: {OUTPUT}")
    print(f"Copy: {OUTPUT_COPY}")
    print(f"Data input belum cocok: {len(unmatched)}")
    print(f"Data diabaikan: {len(ignored)}")
    print(f"Pre ganda: {sum(1 for item in students if len(item.pre_attempts) > 1)}")
    print(f"Post ganda: {sum(1 for item in students if len(item.post_attempts) > 1)}")


if __name__ == "__main__":
    main()
